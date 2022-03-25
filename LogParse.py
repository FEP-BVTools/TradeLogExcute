# -*- coding: utf-8 -*-
"""
Created on Mon Mar 21 08:28:14 2022

@author: ed_liu
"""
import os
import csv 
from openpyxl import load_workbook


class TradeLogParse:
    def GetIPassParseParameter(FileName):
        IPassParameter={}
        
        with open(FileName, newline='') as csvfile:
            rows = csv.DictReader(csvfile)
            #將Value型態設為List
            for row in rows:
                for tital in row.keys():
                    IPassParameter[tital]=[]
        
        #不重複讀檔 迴圈不會跑!!!
        with open(FileName, newline='') as csvfile:
            ros = csv.DictReader(csvfile)
            for ro in ros:
                for tita in ro.keys():            
                    try:
                        IPassParameter[tita].append(eval(ro[tita]))
                    except:
                        print('{}:非數字!!!'.format(tita))
        return IPassParameter
    
    def KRTCParse(IPassParseParameter):
        #取得KRTC資料夾檔案目錄
        KRTCDataList=os.listdir("./Datas/KRTC")
        if len(KRTCDataList)>0:
            for DataFileName in  KRTCDataList:
                if DataFileName.find('DAT')!=-1:
                    #讀取Log資料
                    IPassDataFile=open('./Datas/KRTC/{}'.format(DataFileName))              
                    IPassDatas=IPassDataFile.readlines()
                    IPassDataFile.close()
                    
                    #解析Log資料
                    IPassParseKeyList=[]
                    AfterParse={}
                    AllIPassLogDatas={}
                    # 定義欄位
                    for IPassKey in IPassParseParameter.keys():
                        IPassParseKeyList.append(IPassKey)

                    if len(IPassDatas)>3:
                        
                        OutputFileName=DataFileName.split('.')
                        
                        with open('./Datas/KRTC/{}.csv'.format(OutputFileName[0]), 'w+', newline='') as csvfile:
                            writer = csv.DictWriter(csvfile, fieldnames=IPassParseKeyList)
                            # 寫入第一列的欄位名稱
                            writer.writeheader()
                            writer.writerow(AfterParse)
                            for IPassDataIndex in range(2,len(IPassDatas)-1):                            
                                IPassData=IPassDatas[IPassDataIndex]
                                
                                #以參數進行解析
                                for x in IPassParseParameter.keys():
                                    AfterParse[x]='"{}"'.format(IPassData[IPassParseParameter[x][0]:IPassParseParameter[x][1]])
                                writer.writerow(AfterParse)
                                AllIPassLogDatas[IPassDataIndex]=AfterParse
                    
                    else:
                        print("開班檔:",DataFileName)
        else:
            print("資料夾內無交易Log!!!!!!!!!")
    def KRTCExcelParse():
        # 讀取 Excel 檔案
        wb = load_workbook('公車客運渡輪BVTI解析(new3).xlsx')
        sheet = wb['TxnRecord_解析區']
        
        #取得KRTC資料夾檔案目錄
        KRTCDataList=os.listdir("./Datas/KRTC")
        #確認有檔案
        if len(KRTCDataList)>0:
            #逐個檔案處理
            for DataFileName in  KRTCDataList:
                #僅處理有DAT的檔案
                if DataFileName.find('DAT')!=-1:
                    #讀取Log資料
                    IPassDataFile=open('./Datas/KRTC/{}'.format(DataFileName))              
                    IPassDatas=IPassDataFile.readlines()
                    IPassDataFile.close()
                    
                    # 確認不是開班檔(無交易資料)
                    if len(IPassDatas)>3:
                        #獲取檔名
                        OutputFileName=DataFileName.split('.')
                        #逐筆交易處理
                        
                        for IPassDataIndex in range(2,len(IPassDatas)-1):                            
                            IPassData=IPassDatas[IPassDataIndex]
                            #放入Excel解譯欄位
                            #4代表我的起始位置,不想加變數~~~
                            sheet.cell(row=4+IPassDataIndex, column=1, value=IPassData)
                        # 將修改後的 Excel 檔案進行儲存
                        wb.save('{}.xlsx'.format(OutputFileName[0]))
        
                    
                    else:
                        print("開班檔:",DataFileName)
        else:
            print("資料夾內無交易Log!!!!!!!!!")
        
    def EZCParse():
        
        os.system("move Datas/EZC ./")
        os.system("RENAME EZC logs")
        
        
        os.system("MD output")
        
        os.system("tscc_log_output.exe")
        
        os.system("RENAME logs EZC")
        os.system("xcopy /y output EZC")
        os.system("RD output /Q/S")
        
        os.system("move EZC Datas")
    def ICashParse():
        try:
            ICashExeDirFilePath="ICashExeDir.txt"
            ICashExeDirFile=open(ICashExeDirFilePath,encoding='utf8')
            ICashExeDir=ICashExeDirFile.readline()
            
            os.system("move Datas/ICASH ./")

            os.system("xcopy /y ICASH {}".format(ICashExeDir))
            
            os.system("move ICASH Datas")
            
        except:
            print("ICashExeDir.txt檔不存在或ICash解譯指定位置不存在")
            
        

#確認該資料夾以下所有層是否含有KRTC

if __name__ == '__main__':
    while(1):
        a=input('執行解析')
        print('離開請自行關閉~~')
        if  os.path.exists("Datas"):
            TradeFolderList=os.listdir("./Datas")
    
            for x in TradeFolderList:
    
                if x=="EZC":
                    print("執行{}分析".format(x))
                    TradeLogParse.EZCParse()
    
                
                elif x=="KRTC":
                    print("執行{}分析".format(x))
                    TradeLogParse.KRTCExcelParse()
                    
                elif x=="ICASH":
                    print("執行{}分析".format(x))
                    TradeLogParse.ICashParse()
            print('解析程序結束')
            print('--------------------------------------------------------------')
        else:
            print("無Datas資料夾")

